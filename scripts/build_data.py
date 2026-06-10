# -*- coding: utf-8 -*-
"""把 环境学专业入门词库.xlsx 提取为前端用的 JSON。
输出到 ../data/ :
  simple.json    简易词库
  advanced.json  进阶词库
  directions.json 79个研究方向 -> 单词列表
  meta.json      汇总信息(每本词书的术语类别及数量)
"""
import openpyxl, json, os

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(ROOT, "环境学专业入门词库.xlsx")
DATA = os.path.join(ROOT, "data")
os.makedirs(DATA, exist_ok=True)

COLS = ["word", "reading", "definition", "category", "related", "confusing", "notes"]


def clean(v):
    return str(v).strip() if v is not None else ""


def split_terms(s):
    # 相关词/易混词等用顿号、逗号、空格分隔, 统一切分
    if not s:
        return []
    out = []
    for part in s.replace("、", ",").replace("，", ",").replace("、", ",").split(","):
        p = part.strip()
        if p:
            out.append(p)
    return out


def extract_book(ws):
    rows = ws.iter_rows(values_only=True)
    next(rows)  # header
    items = []
    for r in rows:
        word = clean(r[0]) if len(r) > 0 else ""
        if not word:
            continue
        rec = {
            "id": len(items),
            "word": word,
            "reading": clean(r[1]) if len(r) > 1 else "",
            "definition": clean(r[2]) if len(r) > 2 else "",
            "category": clean(r[3]) if len(r) > 3 else "",
            "related": split_terms(clean(r[4]) if len(r) > 4 else ""),
            "confusing": split_terms(clean(r[5]) if len(r) > 5 else ""),
            "notes": clean(r[6]) if len(r) > 6 else "",
        }
        items.append(rec)
    return items


def categories(items):
    from collections import Counter
    c = Counter(it["category"] or "未分类" for it in items)
    return [{"name": k, "count": v} for k, v in c.most_common()]


def main():
    wb = openpyxl.load_workbook(XLSX)
    simple = extract_book(wb["简易单词书"])
    advanced = extract_book(wb["进阶单词书"])

    # 研究方向 sheet
    ws3 = wb["研究方向及与单词生成"]
    rows = ws3.iter_rows(values_only=True)
    next(rows)
    directions = []
    for r in rows:
        name = clean(r[0]) if len(r) > 0 else ""
        if not name:
            continue
        words = split_terms(clean(r[1]) if len(r) > 1 else "")
        directions.append({"name": name, "words": words})

    meta = {
        "books": {
            "simple": {"label": "简易词库", "count": len(simple), "categories": categories(simple)},
            "advanced": {"label": "进阶词库", "count": len(advanced), "categories": categories(advanced)},
        },
        "directionCount": len(directions),
    }

    def dump_js(key, obj):
        s = json.dumps(obj, ensure_ascii=False, separators=(",", ":"))
        body = "window.__HJ__=window.__HJ__||{};window.__HJ__.%s=%s;" % (key, s)
        open(os.path.join(DATA, key + ".js"), "w", encoding="utf-8").write(body)

    # JS 形式(可在 file:// 下用 <script> 加载, 免 fetch/CORS)
    dump_js("simple", simple)
    dump_js("advanced", advanced)
    dump_js("directions", directions)
    dump_js("meta", meta)

    # 同时保留 JSON (方便其他用途)
    for name, obj in [("simple", simple), ("advanced", advanced),
                      ("directions", directions), ("meta", meta)]:
        json.dump(obj, open(os.path.join(DATA, name + ".json"), "w", encoding="utf-8"),
                  ensure_ascii=False, separators=(",", ":"))

    print("simple:", len(simple), "advanced:", len(advanced), "directions:", len(directions))


if __name__ == "__main__":
    main()
