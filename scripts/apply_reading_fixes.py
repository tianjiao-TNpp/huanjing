# -*- coding: utf-8 -*-
"""把 corrections_auto.json(或指定的修正映射) 应用到 源Excel 的两本词书,
并记录改动日志。之后请再跑 build_data.py 重新生成 data。
用法:
  python scripts/apply_reading_fixes.py                # 用 corrections_auto.json
  python scripts/apply_reading_fixes.py 某映射.json     # 用自定义 {单词:读音}
"""
import json, os, sys
import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(ROOT, "环境学专业入门词库.xlsx")
SHEETS = ["简易单词书", "进阶单词书"]
WORD_COL, READ_COL = 1, 2   # A=单词, B=假名/读音 (1-based)


def main():
    mapf = sys.argv[1] if len(sys.argv) > 1 else os.path.join(ROOT, "corrections_auto.json")
    fixes = json.load(open(mapf, encoding="utf-8"))
    wb = openpyxl.load_workbook(XLSX)
    changed = []
    for sh in SHEETS:
        ws = wb[sh]
        for row in range(2, ws.max_row + 1):
            w = ws.cell(row, WORD_COL).value
            if w is None:
                continue
            w = str(w).strip()
            if w in fixes:
                old = ws.cell(row, READ_COL).value
                new = fixes[w]
                if str(old).strip() != new:
                    ws.cell(row, READ_COL).value = new
                    changed.append({"sheet": sh, "word": w, "old": str(old).replace("\n", " ")[:50], "new": new})
    wb.save(XLSX)
    json.dump(changed, open(os.path.join(ROOT, "读音改动日志.json"), "w", encoding="utf-8"),
              ensure_ascii=False, indent=1)
    print("应用映射词数:", len(fixes), "| 实际改动单元格:", len(changed))


if __name__ == "__main__":
    main()
