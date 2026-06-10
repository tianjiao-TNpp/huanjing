# -*- coding: utf-8 -*-
"""读取用户编辑过的 读音待确认.xlsx:
 - 黄色(FFFF00)行 -> 该单词从两本词书中删除(中文/损坏的重复条目)
 - 其余行 -> 用"你确认填这列"(空则用"建议读音")更新读音
另外对几个客观读音错误做标准化纠正(已在报告中说明)。
直接改源 环境学专业入门词库.xlsx, 并写删除/改动日志。
"""
import json, os
import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
REVIEW = os.path.join(ROOT, "读音待确认.xlsx")
XLSX = os.path.join(ROOT, "环境学专业入门词库.xlsx")
SHEETS = ["简易单词书", "进阶单词书"]

# 客观纠错(标准读音), 已在对话中向用户说明
OVERRIDE = {
    "先進水処理": "せんしんすいしょり",   # 先進=せんしん(非さきしん)
    "水辺": "みずべ",                     # 水辺=みずべ(みなべ非其读音)
    "湿地帯": "しっちたい",               # 湿地=しっち(促音)
    "脱窒素": "だっちっそ",               # 窒素=ちっそ(促音)
    "脱窒素装置": "だっちっそそうち",
}


def main():
    wb = openpyxl.load_workbook(REVIEW)
    ws = wb.active
    delete_words, read_map = set(), {}
    for r in range(2, ws.max_row + 1):
        word = ws.cell(r, 1).value
        if not word:
            continue
        word = str(word).strip()
        fill = ws.cell(r, 1).fill
        is_yellow = (fill and fill.patternType == "solid"
                     and str(getattr(fill.fgColor, "rgb", "")).endswith("FFFF00"))
        if is_yellow:
            delete_words.add(word)
            continue
        confirm = ws.cell(r, 5).value
        rec = ws.cell(r, 3).value
        reading = (str(confirm).strip() if confirm not in (None, "") else str(rec).strip())
        read_map[word] = reading
    read_map.update(OVERRIDE)
    # 删除优先级高于改读音
    for w in delete_words:
        read_map.pop(w, None)

    src = openpyxl.load_workbook(XLSX)
    deleted, changed = [], []
    for sh in SHEETS:
        s = src[sh]
        del_rows = []
        for row in range(2, s.max_row + 1):
            w = s.cell(row, 1).value
            if w is None:
                continue
            w = str(w).strip()
            if w in delete_words:
                del_rows.append(row)
                deleted.append({"sheet": sh, "word": w, "row": row})
            elif w in read_map:
                old = s.cell(row, 2).value
                if str(old).strip() != read_map[w]:
                    s.cell(row, 2).value = read_map[w]
                    changed.append({"sheet": sh, "word": w, "old": str(old)[:30], "new": read_map[w]})
        for row in reversed(del_rows):   # 从下往上删, 保持行号
            s.delete_rows(row, 1)
    src.save(XLSX)

    json.dump({"deleted": deleted, "changed": changed},
              open(os.path.join(ROOT, "读音改动日志2.json"), "w", encoding="utf-8"),
              ensure_ascii=False, indent=1)
    print("删除条目:", len(deleted), "(单词:", sorted(delete_words), ")")
    print("更新读音单元格:", len(changed))


if __name__ == "__main__":
    main()
