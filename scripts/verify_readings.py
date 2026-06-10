# -*- coding: utf-8 -*-
"""校验词库里日语单词的读音(假名)是否正确。
方法: 用 UniDic(形态素解析) 和 pykakasi 两套独立生成读音, 与现存读音比对。
输出 读音校对.xlsx: 只列出"现存读音 与两套工具都不一致 / 或与解析不一致"的待复核条目,
按可信度排序(两套工具一致且都≠现存 => 高可信错误)。
"""
import json, os, sys
import fugashi, pykakasi

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA = os.path.join(ROOT, "data")

tagger = fugashi.Tagger()
kks = pykakasi.kakasi()


def to_hira(s):
    out = []
    for ch in s:
        o = ord(ch)
        if 0x30A1 <= o <= 0x30F6:      # 片假名 -> 平假名
            out.append(chr(o - 0x60))
        else:
            out.append(ch)
    return "".join(out)


def to_kata(s):
    out = []
    for ch in s:
        o = ord(ch)
        if 0x3041 <= o <= 0x3096:      # 平假名 -> 片假名
            out.append(chr(o + 0x60))
        else:
            out.append(ch)
    return "".join(out)


def norm(s):
    """用于比较的归一化: 统一片假名, 去掉长音/分隔/空白等干扰。"""
    s = to_kata((s or "").strip())
    for c in "ー・　  \t‐-―~〜":
        s = s.replace(c, "")
    return s


def uni_reading(w):
    parts = []
    for m in tagger(w):
        k = None
        try:
            k = m.feature.kana
        except Exception:
            k = None
        parts.append(k if k and k != "*" else m.surface)
    return "".join(parts)


def kaka_reading(w):
    return "".join(it["kana"] for it in kks.convert(w))  # 片假名


def has_kanji(s):
    return any(0x4E00 <= ord(c) <= 0x9FFF or 0x3400 <= ord(c) <= 0x4DBF for c in s)


def load_pairs():
    """返回 {word: {"reading": 现存读音, "category":.., "books": set}} ，
    若两本书读音不同则分别记录。键用 (word, reading)。"""
    pairs = {}
    for book in ("simple", "advanced"):
        arr = json.load(open(os.path.join(DATA, book + ".js" if False else book + ".json"), encoding="utf-8"))
        for w in arr:
            word = (w.get("word") or "").strip()
            reading = (w.get("reading") or "").strip()
            if not word:
                continue
            key = (word, reading)
            if key not in pairs:
                pairs[key] = {"word": word, "reading": reading,
                              "category": w.get("category", ""), "books": set()}
            pairs[key]["books"].add(book)
    return pairs


def main():
    pairs = load_pairs()
    rows = []
    for (word, reading), info in pairs.items():
        if not has_kanji(word):
            continue  # 纯假名/外来语(片假名)/英文, 读音≈本身, 跳过
        try:
            u = uni_reading(word)
        except Exception:
            u = ""
        try:
            k = kaka_reading(word)
        except Exception:
            k = ""
        ns, nu, nk = norm(reading), norm(u), norm(k)
        if not ns:
            verdict = "缺读音"
        elif ns == nu or ns == nk:
            continue  # 与某一工具一致, 视为通过
        else:
            verdict = "高可信错误" if (nu and nu == nk) else "待复核"
        suggest = to_hira(u) if u else to_hira(k)
        rows.append({
            "word": word, "cur": reading,
            "unidic": to_hira(u), "kakasi": to_hira(k),
            "verdict": verdict, "suggest": suggest,
            "category": info["category"], "books": "/".join(sorted(info["books"])),
        })

    order = {"高可信错误": 0, "待复核": 1, "缺读音": 2}
    rows.sort(key=lambda r: (order.get(r["verdict"], 9), r["word"]))

    # 写 Excel
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "读音校对"
    headers = ["单词", "现存读音", "UniDic读音", "kakasi读音", "判定", "建议读音(确认后填这个)", "术语类别", "出现于"]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="DDEEE5")
    redfill = PatternFill("solid", fgColor="FFE0E0")
    for r in rows:
        ws.append([r["word"], r["cur"], r["unidic"], r["kakasi"], r["verdict"],
                   r["suggest"], r["category"], r["books"]])
        if r["verdict"] == "高可信错误":
            for c in ws[ws.max_row]:
                c.fill = redfill
    widths = [16, 16, 16, 16, 12, 22, 16, 10]
    for i, wd in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = wd
    ws.freeze_panes = "A2"
    out = os.path.join(ROOT, "读音校对.xlsx")
    wb.save(out)

    hi = sum(1 for r in rows if r["verdict"] == "高可信错误")
    rev = sum(1 for r in rows if r["verdict"] == "待复核")
    miss = sum(1 for r in rows if r["verdict"] == "缺读音")
    print("待校对条目:", len(rows))
    print("  高可信错误:", hi)
    print("  待复核:", rev)
    print("  缺读音:", miss)
    print("输出:", out)
    # 同时存一份 JSON 便于程序查看前若干条
    json.dump(rows[:40], open(os.path.join(ROOT, "_readings_sample.json"), "w", encoding="utf-8"),
              ensure_ascii=False, indent=1)


if __name__ == "__main__":
    main()
