# -*- coding: utf-8 -*-
"""把读音校验结果分成两组:
  corrections_auto.json  : 确定无误, 可自动应用 {word: 正确读音}
  读音待确认.xlsx        : 音训敏感 / 工具分歧 / 词条本身可能损坏 -> 交人最终确认
基于人工已确认的误检模式(塩→えん, 光→こう, 嫌気→けんき, 生分解→せい, 廃紙, 前処理,
水+汉语→すい, 油, 飛灰 等)做保守过滤: 拿不准的一律进"待确认", 不自动改。
"""
import json, os, re, sys
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import verify_readings as v

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


def has_cjk(s):
    return any(0x3400 <= ord(c) <= 0x9FFF for c in s)


def classify(word, stored, u, k, verdict):
    """返回 (bucket, reading, note)。bucket in {auto, user, skip}"""
    # 词条本身含中文简体等非日文字符 -> 解析后仍残留汉字 -> 词条损坏, 交人核对
    if has_cjk(u) or has_cjk(k):
        return ("user", stored, "词条疑似损坏(含非日文字符), 请先核对单词本身")
    # 工具分歧 -> 交人
    if verdict != "高":
        return ("user", u, "两套工具读音不一致, 请确认")

    # ---- 音训敏感词: 用我确认过的规则给"建议", 交人最终拍板 ----
    rec, note = None, None
    if word.endswith("塩") and u.endswith("しお"):
        rec, note = u[:-2] + "えん", "塩=えん(音读)"
    elif "嫌気" in word and "いやけ" in u:
        rec, note = u.replace("いやけ", "けんき"), "嫌気=けんき"
    elif word == "光分解" or ("光" in word and u.startswith("ひかり")):
        rec, note = u.replace("ひかり", "こう"), "光=こう(此处音读)"
    elif "前処理" in word and "まえ" in u:
        rec, note = u.replace("まえしょり", "ぜんしょり"), "前処理=ぜんしょり"
    elif "生分解" in word and "なまぶんかい" in u:
        rec, note = u.replace("なまぶんかい", "せいぶんかい"), "生分解=せいぶんかい"
    elif word == "廃紙":
        rec, note = "はいし", "廃紙=はいし"
    elif word == "飛灰":
        rec, note = stored, "飛灰: ひばい/ひはい 请确认"
    elif "油" in word and "あぶら" in u:
        rec, note = stored, "油: 音读ゆ/训读あぶら 请确认"
    elif word.startswith("水") and u.startswith("みず"):
        rec, note = "すい" + u[2:], "水=すい(漢語复合) 请确认"

    if rec is not None:
        if v.norm(rec) == v.norm(stored):
            return ("skip", stored, note)   # 现存其实已对, 不动
        return ("user", rec, note)

    # ---- 其余: 确定无误, 自动应用 ----
    return ("auto", u, "")


def main():
    pairs = v.load_pairs()
    auto, user = {}, []
    for (word, reading), info in pairs.items():
        if not v.has_kanji(word):
            continue
        u = v.to_hira(v.uni_reading(word))
        k = v.to_hira(v.kaka_reading(word))
        ns, nu, nk = v.norm(reading), v.norm(u), v.norm(k)
        if not ns or ns == nu or ns == nk:
            continue
        verdict = "高" if (nu and nu == nk) else "复核"
        bucket, rd, note = classify(word, reading, u, k, verdict)
        if bucket == "auto":
            # 同一单词若多处, 用同一正确读音
            auto[word] = rd
        elif bucket == "user":
            user.append({"word": word, "cur": reading.replace("\n", " ")[:60],
                         "rec": rd, "note": note, "category": info["category"],
                         "books": "/".join(sorted(info["books"]))})
    # auto 里如果某词又出现在 user(敏感), 以 user 为准, 从 auto 删
    user_words = set(x["word"] for x in user)
    for w in list(auto):
        if w in user_words:
            del auto[w]

    json.dump(auto, open(os.path.join(ROOT, "corrections_auto.json"), "w", encoding="utf-8"),
              ensure_ascii=False, indent=1)

    # 写待确认 Excel
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "待确认读音"
    ws.append(["单词", "现存读音", "建议读音", "原因", "你确认填这列(留空=用建议)", "术语类别", "出现于"])
    for c in ws[1]:
        c.font = Font(bold=True); c.fill = PatternFill("solid", fgColor="DDEEE5")
    user.sort(key=lambda r: (r["note"], r["word"]))
    for r in user:
        ws.append([r["word"], r["cur"], r["rec"], r["note"], "", r["category"], r["books"]])
    for i, wd in enumerate([16, 20, 16, 26, 22, 14, 10], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = wd
    ws.freeze_panes = "A2"
    wb.save(os.path.join(ROOT, "读音待确认.xlsx"))

    print("自动应用(确定):", len(auto))
    print("待人确认:", len(user))
    # 备份一份 auto 的可读样本
    json.dump(dict(list(auto.items())[:50]), open(os.path.join(ROOT, "_auto_sample.json"), "w", encoding="utf-8"),
              ensure_ascii=False, indent=1)


if __name__ == "__main__":
    main()
